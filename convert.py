#!/usr/bin/env python3
"""
Excel → JSON 转换脚本
自动识别两种 Excel 格式并转换为甘特图所需的 JSON 文件

支持格式:
  A) 模板格式（create_template.py 生成）
     - 第1个Sheet「配置」含全局设置 + 项目列表 + 里程碑
     - 后续Sheet: 类型 | 任务名称 | 优先级 | 负责人 | 开始日期 | 结束日期 | 状态
  B) 项目管理格式（用户日常维护的）
     - 每个Sheet就是一个项目（如"强控新西兰"）
     - 第1行: 上线目标 | 630前上线
     - 第2行: 服务商 | 强控
     - 第3行: 表头（阶段 | 事项 | 优先级 | 责任人 | 目标完成时间 | 实际完成时间 | 当前状态 | ...）
     - A列非空行 = 阶段标题，A列空B列有值但无优先级/状态 = 子分组

使用方法:
    cd project-gantt
    python3 convert.py                          # 自动识别 data/ 下的 xlsx
    python3 convert.py data/开仓项目管理.xlsx    # 指定文件
"""
import json
import os
import re
import sys
from datetime import datetime, timedelta

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, "data")

STATUS_MAP = {
    "已完成": "done", "完成": "done", "done": "done", "✅": "done",
    "进行中": "progress", "进行": "progress", "progress": "progress", "🔄": "progress",
    "未开始": "todo", "待开始": "todo", "todo": "todo", "⏳": "todo",
}

# 项目颜色轮替
PROJECT_COLORS = ["#1a73e8", "#e65100", "#2e7d32", "#6a1b9a", "#00838f", "#c62828", "#4527a0", "#ef6c00"]
# 项目国旗 emoji 映射
COUNTRY_FLAGS = {
    "新西兰": "🇳🇿", "意波英荷": "🇮🇹🇵🇱🇬🇧🇳🇱", "意大利": "🇮🇹",
    "智利": "🇨🇱", "以色列": "🇮🇱", "韩国": "🇰🇷", "日本": "🇯🇵",
    "美国": "🇺🇸", "英国": "🇬🇧", "法国": "🇫🇷", "德国": "🇩🇪",
    "西班牙": "🇪🇸", "荷兰": "🇳🇱", "波兰": "🇵🇱", "巴西": "🇧🇷",
    "墨西哥": "🇲🇽", "泰国": "🇹🇭", "越南": "🇻🇳", "马来": "🇲🇾",
    "新加坡": "🇸🇬", "印尼": "🇮🇩", "菲律宾": "🇵🇭", "澳大利亚": "🇦🇺",
}

# Excel 日期序列号基准 (1900-01-01 = 1, 但Excel有1900-02-29 bug)
EXCEL_EPOCH = datetime(1899, 12, 30)


def excel_serial_to_date(serial):
    """将 Excel 日期序列号转为 datetime"""
    try:
        serial_int = int(float(serial))
        if 40000 <= serial_int <= 55000:  # 合理的 Excel 日期范围 (2009-2050)
            return EXCEL_EPOCH + timedelta(days=serial_int)
    except (ValueError, TypeError):
        pass
    return None


def parse_date(value):
    """将各种日期格式转为 YYYY-MM-DD 字符串，无法解析则返回空串"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    text = str(value).strip()
    if not text:
        return ""

    # 尝试 Excel 序列号
    serial_date = excel_serial_to_date(text)
    if serial_date:
        return serial_date.strftime("%Y-%m-%d")

    # 尝试标准日期格式
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    # 尝试 "X月" "X月初" "X月中" "X月底" 等模糊日期（使用当前年份或2026）
    month_match = re.match(r"(\d{1,2})月", text)
    if month_match:
        month = int(month_match.group(1))
        year = 2026  # 默认年份
        if "底" in text or "末" in text:
            day = 28
        elif "中" in text:
            day = 15
        elif "初" in text:
            day = 5
        else:
            day = 15  # 默认月中
        try:
            return f"{year}-{month:02d}-{day:02d}"
        except Exception:
            pass

    return ""


def safe_str(value):
    """安全转字符串，None 返回空串"""
    if value is None:
        return ""
    return str(value).strip()


def clean_owner(text):
    """清理责任人字段：去除@符号，清理多余空格"""
    if not text:
        return ""
    text = text.strip()
    # 去掉开头的 @
    text = re.sub(r"^[@\s]+", "", text)
    # 处理多人：@A @B 或 @A、@B
    text = re.sub(r"\s*[@]\s*", "/", text)
    text = re.sub(r"\s*[、，,]\s*@?\s*", "/", text)
    # 清理多余空格
    text = re.sub(r"\s+", " ", text).strip()
    return text


def get_country_flag(sheet_name):
    """根据 Sheet 名称获取国旗 emoji"""
    for country, flag in COUNTRY_FLAGS.items():
        if country in sheet_name:
            return flag
    return "🏳️"


def extract_project_name(sheet_name):
    """从 Sheet 名称中提取项目短名（去掉"强控""自邮"等前缀）"""
    name = sheet_name
    for prefix in ["强控", "自邮", "平台", "项目"]:
        name = name.replace(prefix, "")
    return name.strip()


def is_phase_row(col_a_value):
    """判断是否是阶段行（A列有值且以数字+、开头）"""
    text = safe_str(col_a_value)
    if not text:
        return False
    # 匹配 "1、" "2、" "1." "第一阶段" 等
    if re.match(r"^\d+[、.．]", text):
        return True
    if re.match(r"^第[一二三四五六七八九十]+", text):
        return True
    return False


def is_subgroup_row(row_values):
    """判断是否是子分组行（A列空，B列有值，但没有优先级和状态）"""
    col_a = safe_str(row_values[0])
    col_b = safe_str(row_values[1])
    col_c = safe_str(row_values[2])  # 优先级
    col_g = safe_str(row_values[6]) if len(row_values) > 6 else ""  # 当前状态

    if col_a:
        return False
    if not col_b:
        return False
    # 子分组一般无优先级且无状态
    has_priority = col_c.upper() in ("P0", "P1", "P2")
    has_status = col_g in STATUS_MAP
    if not has_priority and not has_status:
        return True
    return False


def detect_format(workbook):
    """自动检测 Excel 格式：'template' 或 'project_mgmt'"""
    first_sheet = workbook.worksheets[0]
    cell_a1 = safe_str(first_sheet.cell(row=1, column=1).value)

    # 模板格式第一个 Sheet 标题为「配置」，A1 包含"甘特图"或"全局配置"
    if first_sheet.title == "配置" or "甘特图" in cell_a1 or "全局配置" in cell_a1:
        return "template"

    # 项目管理格式第一行 A1 是"上线目标"
    if "上线目标" in cell_a1 or "上线" in cell_a1:
        return "project_mgmt"

    # 检查是否有类似"阶段|事项"的表头
    for row_idx in range(1, 5):
        a_val = safe_str(first_sheet.cell(row=row_idx, column=1).value)
        if "阶段" in a_val:
            return "project_mgmt"

    # 默认当作模板格式
    return "template"


# ==================== 模板格式解析 ====================

def parse_template_config_sheet(ws):
    """解析模板格式的配置 Sheet"""
    config = {
        "title": "",
        "subtitle": "",
        "timeRange": {"start": "", "end": ""},
        "milestones": [],
        "projects": [],
    }

    key_map = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=False):
        key = safe_str(row[0].value)
        if key:
            key_map[key] = row[1].value

    config["title"] = safe_str(key_map.get("标题", ""))
    config["subtitle"] = safe_str(key_map.get("副标题", ""))
    config["timeRange"]["start"] = parse_date(key_map.get("时间范围开始", ""))
    config["timeRange"]["end"] = parse_date(key_map.get("时间范围结束", ""))

    project_header_row = None
    milestone_header_row = None
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=False):
        text = safe_str(row[0].value)
        if "项目列表" in text:
            project_header_row = row[0].row + 1
        elif "里程碑" in text:
            milestone_header_row = row[0].row + 1

    if project_header_row:
        for row in ws.iter_rows(min_row=project_header_row + 1, max_col=6, values_only=True):
            key = safe_str(row[0])
            if not key:
                break
            config["projects"].append({
                "key": key, "file": key + ".json",
                "shortName": safe_str(row[2]), "sheetName": safe_str(row[1]),
                "color": safe_str(row[3]),
                "deadline": parse_date(row[4]), "deadlineLabel": safe_str(row[5]),
            })

    if milestone_header_row:
        for row in ws.iter_rows(min_row=milestone_header_row + 1, max_col=4, values_only=True):
            label = safe_str(row[1])
            if not label:
                break
            config["milestones"].append({
                "date": parse_date(row[0]), "label": label,
                "color": safe_str(row[2]),
                "projects": [k.strip() for k in safe_str(row[3]).split(",") if k.strip()],
            })

    return config


def parse_template_project_sheet(ws):
    """解析模板格式的项目 Sheet"""
    tasks = []
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        row_type = safe_str(row[0]).lower()
        name = safe_str(row[1])
        if not name:
            continue
        if row_type in ("阶段", "phase"):
            tasks.append({"phase": name, "start": parse_date(row[4]), "end": parse_date(row[5]), "status": "phase"})
        else:
            tasks.append({
                "task": name, "p": safe_str(row[2]).upper() if safe_str(row[2]).upper() in ("P0", "P1", "P2") else "",
                "owner": safe_str(row[3]), "start": parse_date(row[4]), "end": parse_date(row[5]),
                "status": STATUS_MAP.get(safe_str(row[6]).lower(), "todo"),
            })
    return tasks


# ==================== 项目管理格式解析 ====================

def parse_project_mgmt_sheet(ws):
    """
    解析项目管理格式的 Sheet
    返回 (meta_info, tasks)
    meta_info: { 'deadline_label': '630', 'provider': '强控' }
    """
    meta = {"deadline_label": "", "provider": ""}
    tasks = []

    # 读取前两行元信息
    row1_a = safe_str(ws.cell(row=1, column=1).value)
    row1_b = safe_str(ws.cell(row=1, column=2).value)
    row2_a = safe_str(ws.cell(row=2, column=1).value)
    row2_b = safe_str(ws.cell(row=2, column=2).value)

    if "上线" in row1_a:
        meta["deadline_label"] = row1_b.replace("前上线", "").replace("上线", "").strip()
    if "服务商" in row2_a:
        meta["provider"] = row2_b

    # 从第4行开始解析数据（第3行是表头）
    current_phase = None

    for row_idx in range(4, ws.max_row + 1):
        row_values = []
        for col_idx in range(1, ws.max_column + 1):
            row_values.append(ws.cell(row=row_idx, column=col_idx).value)

        # 至少需要2列数据
        while len(row_values) < 9:
            row_values.append(None)

        col_a = safe_str(row_values[0])  # 阶段
        col_b = safe_str(row_values[1])  # 事项
        col_c = safe_str(row_values[2])  # 优先级
        col_d = safe_str(row_values[3])  # 责任人
        col_e = row_values[4]            # 目标完成时间（保留原始类型）
        col_f = row_values[5]            # 实际完成时间
        col_g = safe_str(row_values[6])  # 当前状态
        col_h = safe_str(row_values[7])  # 风险评估
        col_i = safe_str(row_values[8]) if len(row_values) > 8 else ""  # 项目进展备注

        # 跳过完全空行
        if not col_a and not col_b:
            continue

        # 阶段行
        if is_phase_row(col_a):
            current_phase = col_a
            # 阶段行的 B 列是该阶段的第一个任务
            tasks.append({
                "phase": col_a,
                "start": "",
                "end": "",
                "status": "phase",
            })
            # 如果 B 列也有任务名，当作任务处理
            if col_b:
                target_date = parse_date(col_e)
                actual_date = parse_date(col_f)
                status = STATUS_MAP.get(col_g.lower(), "todo") if col_g else "todo"
                priority = col_c.upper() if col_c.upper() in ("P0", "P1", "P2") else ""
                task_entry = {
                    "task": col_b.strip(),
                    "p": priority,
                    "owner": clean_owner(col_d),
                    "start": actual_date or target_date,
                    "end": target_date or actual_date,
                    "status": status,
                    "risk": col_h,
                    "remark": col_i,
                }
                if task_entry["start"] or task_entry["end"]:
                    tasks.append(task_entry)
                elif status != "todo":
                    tasks.append(task_entry)
            continue

        # A列为空的行
        if not col_a and col_b:
            # 判断是子分组还是任务
            if is_subgroup_row(row_values):
                # 子分组标题不作为独立任务，跳过
                continue

            # 任务行
            target_date = parse_date(col_e)
            actual_date = parse_date(col_f)
            status = STATUS_MAP.get(col_g.lower(), "todo") if col_g else "todo"
            priority = col_c.upper() if col_c.upper() in ("P0", "P1", "P2") else ""

            task_name = col_b.strip()
            # 去掉任务名前面的缩进空格
            task_name = re.sub(r"^\s+", "", task_name)

            task_entry = {
                "task": task_name,
                "p": priority,
                "owner": clean_owner(col_d),
                "start": actual_date or target_date,
                "end": target_date or actual_date,
                "status": status,
                "risk": col_h,
                "remark": col_i,
            }
            # 只要有名称就加入（日期可以为空）
            tasks.append(task_entry)

    return meta, tasks


def estimate_deadline_date(deadline_label, year=2026):
    """根据截止标签（如 '630'、'730'）推算截止日期"""
    match = re.match(r"(\d{1,2})(\d{2})", deadline_label)
    if match:
        month = int(match.group(1))
        day = int(match.group(2))
        try:
            return f"{year}-{month:02d}-{day:02d}"
        except Exception:
            pass
    return ""


def compute_phase_dates(tasks):
    """根据任务的日期范围回填阶段的起止日期"""
    phase_indices = []
    for i, task in enumerate(tasks):
        if task.get("status") == "phase":
            phase_indices.append(i)

    for idx, phase_idx in enumerate(phase_indices):
        # 收集该阶段下所有任务的日期
        next_phase_idx = phase_indices[idx + 1] if idx + 1 < len(phase_indices) else len(tasks)
        min_start = None
        max_end = None
        for j in range(phase_idx + 1, next_phase_idx):
            task = tasks[j]
            if task.get("start"):
                try:
                    date = datetime.strptime(task["start"], "%Y-%m-%d")
                    if min_start is None or date < min_start:
                        min_start = date
                except ValueError:
                    pass
            if task.get("end"):
                try:
                    date = datetime.strptime(task["end"], "%Y-%m-%d")
                    if max_end is None or date > max_end:
                        max_end = date
                except ValueError:
                    pass

        if min_start:
            tasks[phase_idx]["start"] = min_start.strftime("%Y-%m-%d")
        if max_end:
            tasks[phase_idx]["end"] = max_end.strftime("%Y-%m-%d")


def expand_task_dates(tasks):
    """
    智能扩展任务日期，让每个任务都有合理的时间跨度（而非一天的细线）。
    规则：
    - 如果 start == end（只有一个日期），根据优先级推算跨度
    - 如果只有 end 没有 start，往前推
    - 如果只有 start 没有 end，往后推
    - 已完成的任务如果有 actual_date，用它来确定 end
    """
    DURATION_BY_PRIORITY = {
        "P0": 14,  # P0 任务通常跨度较长
        "P1": 10,
        "P2": 7,
        "": 7,     # 默认
    }

    for task in tasks:
        if task.get("status") == "phase":
            continue

        start_str = task.get("start", "")
        end_str = task.get("end", "")

        if not start_str and not end_str:
            continue

        priority = task.get("p", "")
        duration_days = DURATION_BY_PRIORITY.get(priority, 7)

        try:
            if start_str and end_str:
                start_dt = datetime.strptime(start_str, "%Y-%m-%d")
                end_dt = datetime.strptime(end_str, "%Y-%m-%d")
                # 如果 start == end 或 end < start，扩展日期
                if (end_dt - start_dt).days < 3:
                    # 以 end 为截止日期，往前推
                    task["start"] = (end_dt - timedelta(days=duration_days)).strftime("%Y-%m-%d")
                    task["end"] = end_dt.strftime("%Y-%m-%d")
            elif end_str and not start_str:
                end_dt = datetime.strptime(end_str, "%Y-%m-%d")
                task["start"] = (end_dt - timedelta(days=duration_days)).strftime("%Y-%m-%d")
            elif start_str and not end_str:
                start_dt = datetime.strptime(start_str, "%Y-%m-%d")
                task["end"] = (start_dt + timedelta(days=duration_days)).strftime("%Y-%m-%d")
        except ValueError:
            continue


def find_excel_file():
    """在 data/ 目录下查找 Excel 文件"""
    xlsx_files = [f for f in os.listdir(DATA_DIR) if f.endswith(".xlsx") and not f.startswith("~$")]
    if not xlsx_files:
        return None
    # 优先使用用户自己的文件（非模板）
    for f in xlsx_files:
        if f != "项目计划.xlsx":
            return os.path.join(DATA_DIR, f)
    return os.path.join(DATA_DIR, xlsx_files[0])


def main():
    # 确定 Excel 文件路径
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
        if not os.path.isabs(excel_path):
            excel_path = os.path.join(SCRIPT_DIR, excel_path)
    else:
        excel_path = find_excel_file()

    if not excel_path or not os.path.exists(excel_path):
        print("❌ 未找到 Excel 文件")
        print(f"   请将 Excel 文件放到 {DATA_DIR}/ 目录下")
        print(f"   或指定路径: python3 convert.py <excel文件路径>")
        sys.exit(1)

    print(f"📖 读取 Excel: {excel_path}")
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    print(f"   Sheet列表: {workbook.sheetnames}")

    # 自动检测格式
    file_format = detect_format(workbook)
    print(f"   检测到格式: {'模板格式' if file_format == 'template' else '项目管理格式'}")
    print()

    if file_format == "template":
        run_template_mode(workbook)
    else:
        run_project_mgmt_mode(workbook)


def run_template_mode(workbook):
    """模板格式处理流程"""
    config = parse_template_config_sheet(workbook.worksheets[0])
    print(f"✅ 配置解析完成: {config['title']}")
    print(f"   时间范围: {config['timeRange']['start']} → {config['timeRange']['end']}")

    sheet_to_project = {p["sheetName"]: p for p in config["projects"]}
    generated_files = []

    for sheet_name in workbook.sheetnames[1:]:
        if sheet_name not in sheet_to_project:
            print(f"⚠️  Sheet「{sheet_name}」未在配置中注册，跳过")
            continue
        proj_config = sheet_to_project[sheet_name]
        tasks = parse_template_project_sheet(workbook[sheet_name])
        project_data = {
            "name": proj_config["shortName"], "color": proj_config["color"],
            "deadline": proj_config["deadline"], "deadlineLabel": proj_config["deadlineLabel"],
            "tasks": tasks,
        }
        output_file = os.path.join(DATA_DIR, proj_config["file"])
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(project_data, f, ensure_ascii=False, indent=2)
        phase_count = sum(1 for t in tasks if t.get("phase"))
        task_count = sum(1 for t in tasks if t.get("task"))
        generated_files.append(proj_config["file"])
        print(f"✅ {proj_config['shortName']}: {phase_count}个阶段, {task_count}个任务 → {proj_config['file']}")

    config_output = {
        "title": config["title"], "subtitle": config["subtitle"],
        "timeRange": config["timeRange"], "milestones": config["milestones"],
        "projects": [{"key": p["key"], "file": p["file"], "shortName": p["shortName"]} for p in config["projects"]],
    }
    write_config_json(config_output, generated_files)


def run_project_mgmt_mode(workbook):
    """项目管理格式处理流程"""
    projects_config = []
    generated_files = []
    all_dates = []
    milestones_map = {}  # deadline_label -> [project_keys]

    # 跳过非项目 Sheet（如"开仓填写信息"等）
    skip_keywords = ["填写", "信息", "说明", "模板", "汇总", "统计"]

    for idx, sheet_name in enumerate(workbook.sheetnames):
        # 跳过非项目 Sheet
        if any(kw in sheet_name for kw in skip_keywords):
            print(f"⏭️  跳过非项目Sheet: {sheet_name}")
            continue

        ws = workbook[sheet_name]
        # 检查是否为项目管理格式（第1行有"上线目标"或第3行有"阶段"）
        row1_a = safe_str(ws.cell(row=1, column=1).value)
        row3_a = safe_str(ws.cell(row=3, column=1).value)
        if "上线" not in row1_a and "阶段" not in row3_a:
            print(f"⏭️  跳过非项目Sheet: {sheet_name}")
            continue

        meta, tasks = parse_project_mgmt_sheet(ws)
        compute_phase_dates(tasks)
        expand_task_dates(tasks)
        # 扩展日期后重新计算阶段日期
        compute_phase_dates(tasks)

        # 生成项目 key
        project_name = extract_project_name(sheet_name)
        project_key = f"proj{len(projects_config) + 1}"
        flag = get_country_flag(sheet_name)
        short_name = f"{flag} {project_name}"
        color = PROJECT_COLORS[len(projects_config) % len(PROJECT_COLORS)]

        # 推算截止日期
        deadline_label = meta.get("deadline_label", "")
        deadline_date = estimate_deadline_date(deadline_label)

        # 收集所有日期用于确定时间范围
        for task in tasks:
            for date_field in ("start", "end"):
                date_str = task.get(date_field, "")
                if date_str:
                    try:
                        all_dates.append(datetime.strptime(date_str, "%Y-%m-%d"))
                    except ValueError:
                        pass

        # 收集里程碑
        if deadline_label:
            if deadline_label not in milestones_map:
                milestones_map[deadline_label] = {"date": deadline_date, "keys": []}
            milestones_map[deadline_label]["keys"].append(project_key)

        project_data = {
            "name": short_name,
            "color": color,
            "deadline": deadline_date,
            "deadlineLabel": deadline_label,
            "tasks": tasks,
        }

        output_file = project_key + ".json"
        output_path = os.path.join(DATA_DIR, output_file)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(project_data, f, ensure_ascii=False, indent=2)

        phase_count = sum(1 for t in tasks if t.get("phase"))
        task_count = sum(1 for t in tasks if t.get("task"))
        generated_files.append(output_file)
        projects_config.append({
            "key": project_key, "file": output_file, "shortName": short_name,
        })
        print(f"✅ {short_name}: {phase_count}个阶段, {task_count}个任务 → {output_file}")

    # 自动确定时间范围
    if all_dates:
        time_start = min(all_dates).replace(day=1)  # 月初
        time_end = max(all_dates) + timedelta(days=30)
        time_end = time_end.replace(day=1) + timedelta(days=31)
        time_end = time_end.replace(day=1) - timedelta(days=1)  # 月末
    else:
        time_start = datetime(2026, 4, 1)
        time_end = datetime(2026, 7, 31)

    # 构建里程碑
    milestones = []
    milestone_colors = ["#ff6f00", "#d32f2f", "#1a73e8", "#2e7d32", "#6a1b9a"]
    for i, (label, info) in enumerate(milestones_map.items()):
        milestones.append({
            "date": info["date"],
            "label": label + "上线",
            "color": milestone_colors[i % len(milestone_colors)],
            "projects": info["keys"],
        })

    config_output = {
        "title": "项目管理综合甘特图",
        "subtitle": "",
        "timeRange": {
            "start": time_start.strftime("%Y-%m-%d"),
            "end": time_end.strftime("%Y-%m-%d"),
        },
        "milestones": milestones,
        "projects": projects_config,
    }
    write_config_json(config_output, generated_files)


def write_config_json(config, generated_files):
    """写入 config.json 并输出汇总"""
    config_path = os.path.join(DATA_DIR, "config.json")
    with open(config_path, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    generated_files.append("config.json")

    print(f"\n🎉 转换完成！共生成 {len(generated_files)} 个文件:")
    for filename in generated_files:
        print(f"   📄 data/{filename}")
    print(f"\n💡 刷新浏览器 http://localhost:8080 即可查看更新后的甘特图")


if __name__ == "__main__":
    main()
